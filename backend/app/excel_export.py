from decimal import Decimal
from io import BytesIO
from typing import Callable, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import models

FORMATO_DATA = "DD/MM/YYYY"
FORMATO_HORA = "H:MM"
FORMATO_MOEDA = '"R$" #,##0.00;-"R$" #,##0.00'

COR_CABECALHO = "1E293B"  # slate-900, mesma cor do header do painel
COR_TEXTO_CABECALHO = "FFFFFF"
COR_BORDA = "CBD5E1"  # slate-300


class Coluna:
    def __init__(self, header: str, valor: Callable[[models.OrdemServico], object], tipo: str, largura: int):
        self.header = header
        self.valor = valor
        self.tipo = tipo
        self.largura = largura


def _custo_total(servico: models.OrdemServico) -> Decimal:
    extras = sum(
        (Decimal(str(c.get("valor", 0))) for c in (servico.custos_adicionais or [])),
        Decimal("0"),
    )
    return Decimal(str(servico.custo_operacional or 0)) + extras


def _lucro(servico: models.OrdemServico) -> Decimal:
    return Decimal(str(servico.valor_cobrado or 0)) - _custo_total(servico)


# 🔧 Mesma ideia do MAPEAMENTO_COLUNAS_QUINTOANDAR do frontend: única lista
# que precisa mudar se o QuintoAndar alterar nome/ordem/quantidade de colunas.
MAPEAMENTO_QUINTO_ANDAR: List[Coluna] = [
    Coluna("Data", lambda s: s.data_horario.date(), "data", 12),
    Coluna("Endereço", lambda s: s.endereco, "texto", 60),
    Coluna("Horário", lambda s: s.data_horario.strftime("%H:%M"), "hora", 10),
    Coluna("Contrato", lambda s: s.numero_contrato or "", "texto", 22),
    # A pedido: a coluna "CUSTO" do modelo QuintoAndar traz o valor cobrado do
    # cliente (não o custo operacional) — nome do cabeçalho não muda.
    Coluna("CUSTO", lambda s: Decimal(str(s.valor_cobrado or 0)), "moeda", 18),
    # Único e exclusivamente o texto puro salvo em observacoes — nada de
    # custos adicionais/efetivos misturados aqui (isso já tem colunas
    # próprias no Relatório Geral).
    Coluna("Obs.", lambda s: s.observacoes or None, "texto", 60),
]

MAPEAMENTO_GERAL: List[Coluna] = [
    Coluna("Data", lambda s: s.data_horario.date(), "data", 12),
    Coluna("Horário", lambda s: s.data_horario.strftime("%H:%M"), "hora", 10),
    Coluna("Tipo", lambda s: "QuintoAndar" if s.e_quinto_andar else "Particular", "texto", 14),
    Coluna("Cliente", lambda s: s.nome_cliente, "texto", 24),
    Coluna("Contrato", lambda s: s.numero_contrato or "", "texto", 14),
    Coluna("Endereço", lambda s: s.endereco, "texto", 50),
    Coluna("Status", lambda s: s.status.value, "texto", 14),
    Coluna("Efetivos", lambda s: s.efetivos_nomes or None, "texto", 28),
    Coluna("Valor Cobrado", lambda s: Decimal(str(s.valor_cobrado or 0)), "moeda", 16),
    Coluna("Custo Operacional", lambda s: Decimal(str(s.custo_operacional or 0)), "moeda", 18),
    Coluna(
        "Custos Adicionais",
        lambda s: sum((Decimal(str(c.get("valor", 0))) for c in (s.custos_adicionais or [])), Decimal("0")),
        "moeda",
        18,
    ),
    Coluna("Custo Total", lambda s: _custo_total(s), "moeda", 16),
    Coluna("Lucro", lambda s: _lucro(s), "moeda", 16),
    # Última coluna: só o texto de "Observações da execução" — sem efetivos
    # nem detalhamento de custos extras (esses já têm colunas próprias).
    Coluna("Observações", lambda s: s.observacoes or None, "texto", 50),
]


def _escrever_celula(ws: Worksheet, linha: int, coluna: int, valor, tipo: str):
    celula = ws.cell(row=linha, column=coluna, value=valor)
    if tipo == "data":
        celula.number_format = FORMATO_DATA
    elif tipo == "hora":
        celula.number_format = FORMATO_HORA
    elif tipo == "moeda":
        celula.number_format = FORMATO_MOEDA
    return celula


def _construir_planilha(
    nome_aba: str,
    servicos: Sequence[models.OrdemServico],
    mapeamento: List[Coluna],
    linha_inicial_dados: int,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = nome_aba

    fonte_cabecalho = Font(name="Poppins", size=11, bold=True, color=COR_TEXTO_CABECALHO)
    preenchimento_cabecalho = PatternFill(start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid")
    borda_fina = Border(
        left=Side(style="thin", color=COR_BORDA),
        right=Side(style="thin", color=COR_BORDA),
        top=Side(style="thin", color=COR_BORDA),
        bottom=Side(style="thin", color=COR_BORDA),
    )

    for c, col in enumerate(mapeamento, start=1):
        celula = ws.cell(row=1, column=c, value=col.header)
        celula.font = fonte_cabecalho
        celula.fill = preenchimento_cabecalho
        celula.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = col.largura
    ws.row_dimensions[1].height = 22

    for r, servico in enumerate(servicos):
        linha = linha_inicial_dados + r
        for c, col in enumerate(mapeamento, start=1):
            celula = _escrever_celula(ws, linha, c, col.valor(servico), col.tipo)
            celula.border = borda_fina
            celula.font = Font(name="Poppins", size=10)
            celula.alignment = Alignment(horizontal="left", vertical="center", wrap_text=col.tipo == "texto")

    if servicos:
        _escrever_linha_total(ws, mapeamento, linha_inicial_dados, len(servicos), borda_fina)

    ws.freeze_panes = f"A{linha_inicial_dados}"
    return wb


def _escrever_linha_total(
    ws: Worksheet,
    mapeamento: List[Coluna],
    linha_inicial_dados: int,
    quantidade_linhas: int,
    borda_fina: Border,
) -> None:
    """Linha final com =SOMA(...) nas colunas financeiras ('moeda'), em negrito
    e com preenchimento de destaque — mesmas bordas/alinhamento das demais."""
    primeira_linha = linha_inicial_dados
    ultima_linha = linha_inicial_dados + quantidade_linhas - 1
    linha_total = ultima_linha + 1

    fonte_total = Font(name="Poppins", size=10, bold=True)
    preenchimento_total = PatternFill(start_color=COR_BORDA, end_color=COR_BORDA, fill_type="solid")

    for c, col in enumerate(mapeamento, start=1):
        if c == 1:
            valor, tipo = "TOTAL", "texto"
        elif col.tipo == "moeda":
            letra = get_column_letter(c)
            valor, tipo = f"=SUM({letra}{primeira_linha}:{letra}{ultima_linha})", "moeda"
        else:
            valor, tipo = None, col.tipo

        celula = _escrever_celula(ws, linha_total, c, valor, tipo)
        celula.font = fonte_total
        celula.fill = preenchimento_total
        celula.border = borda_fina
        celula.alignment = Alignment(horizontal="left", vertical="center")


def gerar_relatorio_quinto_andar(servicos: Sequence[models.OrdemServico]) -> BytesIO:
    # LINHA_INICIAL_DADOS = 3, igual ao frontend: cabeçalho na linha 1,
    # linha 2 em branco (espaço reservado pelo modelo oficial do QuintoAndar).
    wb = _construir_planilha("Planilha1", servicos, MAPEAMENTO_QUINTO_ANDAR, linha_inicial_dados=3)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def gerar_relatorio_geral(servicos: Sequence[models.OrdemServico]) -> BytesIO:
    wb = _construir_planilha("Relatório Geral", servicos, MAPEAMENTO_GERAL, linha_inicial_dados=2)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
